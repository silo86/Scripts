#!/usr/bin/env python
# coding: utf-8

# In[11]:


from sqlalchemy import create_engine
import psycopg2
import pandas as pd
import datetime
import os
import glob
os.path.abspath(os.getcwd())
import sys
import errno
from pathlib import Path


# In[12]:


try:
    os.mkdir('output')
except OSError as e:
    if e.errno != errno.EEXIST:
        raise


# In[3]:


export = os.path.abspath(os.getcwd())
dataset_export = os.path.abspath(os.getcwd())
path = os.path.abspath(os.getcwd())
ruta_archivo = os.path.join(os.path.abspath(os.getcwd()),"output/")


# In[4]:


#172.20.0.154
#datos

hostname = os.getenv("ddedbs_ip")
username = os.getenv("ddedbs_name")
password = os.getenv("ddedbs_pass")


# In[13]:


con = psycopg2.connect(host=hostname, user=username, password=password, dbname='saeciv')
lista_db = 'SELECT datname FROM pg_database WHERE datistemplate = false'
bases = pd.read_sql(lista_db,con)


# In[14]:


bases


# In[10]:


csv = pd.read_csv('consultas.csv')
csv


# In[7]:


def doQuery( query, db, nombre ) :
    print("Iniciando lectura de {0} en base de datos {1}\n".format(nombre, db))
    print("{0}\n".format(datetime.datetime.now()))
    dfs = pd.DataFrame()

    for i, fila in bases.iterrows():
        #si pongo el in me toma cjm cjc y calcula 2 veces
        if str(db) == fila.datname:
            #print("Iniciando lectura de {0} en base de datos {1}\n".format(nombre, fila.datname))
            print(f'Iniciando lectura de {nombre} en base de datos {fila.datname}', end=' ')
            con_b = psycopg2.connect(host=hostname, user=username, password=password, dbname = fila.datname)
            df = pd.read_sql(query,con_b)
            dfs = dfs.append(df)
            con_b.close()
           
    
    dfs.to_csv(ruta_archivo+nombre+'.csv', sep=";", encoding='utf-8')
    print("Finalizado: {0}\n".format(datetime.datetime.now()))
    con.close()


# In[8]:


arch = pd.read_csv(path+ '\\'+'consultas.csv', delimiter=";" )
for i, fila in arch.iterrows():
    doQuery( fila['query'], fila['db'], fila['nombre'] )



os.chdir(ruta_archivo)
extension = 'csv'
all_filenames = [i for i in glob.glob('*.{}'.format(extension))]


#combine all files in the list
merged = pd.concat([pd.read_csv(f,error_bad_lines = False) for f in all_filenames ])
#export to csv
merged.to_csv("merged.csv", index=False, encoding='utf-8')


read_file = pd.read_csv ('merged.csv')

merged.to_excel ('merged.xlsx', index = None, header=True)

from openpyxl import load_workbook
wb2 = load_workbook('merged.xlsx')
ws=wb2.create_sheet('metadata')



from openpyxl.utils.dataframe import dataframe_to_rows




for r in dataframe_to_rows(arch, index=False, header=False):
    ws.append(r)
os.chdir("..")
wb2.save('merged.xlsx')


# In[ ]:




