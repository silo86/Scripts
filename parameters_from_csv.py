#!/usr/bin/env python
# coding: utf-8

from sqlalchemy import create_engine
import psycopg2
import pandas as pd
import datetime
import os
import glob
os.path.abspath(os.getcwd())
import sys
import errno
from pathlib import Path


try:
    os.mkdir('output')
except OSError as e:
    if e.errno != errno.EEXIST:
        raise

export = os.path.abspath(os.getcwd())
dataset_export = os.path.abspath(os.getcwd())
path = os.path.abspath(os.getcwd())
ruta_archivo = os.path.join(os.path.abspath(os.getcwd()),"output/")

#data

hostname = os.getenv("ddedbs_ip")
username = os.getenv("ddedbs_name")
password = os.getenv("ddedbs_pass")

#conection
con = psycopg2.connect(host=hostname, user=username, password=password, dbname='saeciv')
lista_db = 'SELECT datname FROM pg_database WHERE datistemplate = false'
bases = pd.read_sql(lista_db,con)
bases = [base for base in bases.datname if base.startswith('sae') and not base.startswith('saemed') and not base.startswith('saeoga') and base not in ('saejes','saesup','saepjt','saeori')]
#reading csv
csv = pd.read_csv('datos.csv',encoding = 'latin-1')
csv


def doQuery( db, unidad, fecha_implementacion) :
    query = f'''
    select 
db,
unidad,
left(dia,4) as anio,
case when (right(dia,4) between '0101' and '0430') then '1' 
     when (right(dia,4) between '0501' and '0831') then '2'
	 else '3'
	 end 
	 as cuatrimestre,
tipo_escrito,
estado,
sistema,
sum(total) as total from(

		SELECT 
		current_database()AS db,
		j.dscr as unidad,
		febo as dia,
		t.dscr tipo_escrito,
		e.dscr as estado,
		case when (j.dscr = '{unidad}' and febo < '{fecha_implementacion}') then 'lex' else 'sae' end as sistema,
		count (*) as total
		FROM public."HIST" AS h 
		left join public."TIES" AS T on h.tiesid = t.tiesid 
		left join public."ESTA" AS e on e.estaid = h.estaid 
		left join public."JUZG" AS j on h.juzgid = j.juzgid
		where t.dscr ilike 'Escritos Ingre%dos%'
		and left(h.febo,4) >= '2019'
		and j.dscr not ilike '%fisc%'
		and j.dscr = '{unidad}'
		group by 
		current_database(), j.dscr, febo, t.dscr,e.dscr,
		case when (j.dscr = '{unidad}' and febo < '{fecha_implementacion}') then 'lex' else 'sae' end
		order by 1,2,3,4,5
	) escritos_por_fecha_implementacion
group by 
db,
unidad,
left(dia,4),
case when (right(dia,4) between '0101' and '0430') then '1' when (right(dia,4) between '0501' and '0831') then '2'else '3' end, 
tipo_escrito,
estado,
sistema
    '''
    print(f"Iniciando lectura de {unidad} en base de datos {db}\n")
    print("{0}\n".format(datetime.datetime.now()))
    
    for fila in bases:
        if str(db) == fila:
            con_b = psycopg2.connect(host=hostname, user=username, password=password, dbname = fila)
            query = query.format(unidad,fecha_implementacion)
            df = pd.read_sql(query,con_b)
            con_b.close()
    df.to_csv(ruta_archivo+unidad+'.csv', sep=";", encoding='utf-8')
    print("Finalizado: {0}\n".format(datetime.datetime.now()))
    con.close()


arch = pd.read_csv(path+ '\\'+'datos.csv', delimiter=";",encoding = 'latin-1' )
for i, fila in arch.iterrows():
    doQuery( fila['db'], fila['unidad'], fila['fecha_implementacion'] )

os.chdir(ruta_archivo)
extension = 'csv'
all_filenames = [i for i in glob.glob('*.{}'.format(extension))]

#combine all files in the list
merged = pd.concat([pd.read_csv(f,error_bad_lines = False) for f in all_filenames ])
#export to csv
merged.to_csv("merged.csv", index=False, encoding='utf-8')

read_file = pd.read_csv ('merged.csv')
merged.to_excel ('merged.xlsx', index = None, header=True)

from openpyxl import load_workbook
wb2 = load_workbook('merged.xlsx')
ws=wb2.create_sheet('metadata')

from openpyxl.utils.dataframe import dataframe_to_rows

for r in dataframe_to_rows(arch, index=False, header=False):
    ws.append(r)
os.chdir("..")
wb2.save('merged.xlsx')